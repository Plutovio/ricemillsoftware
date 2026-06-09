import io
import re
from datetime import date, timedelta
from decimal import Decimal

from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook, load_workbook

from .models import BankGuarantee, DropdownOption
from .serializers import BankGuaranteeSerializer, DropdownOptionSerializer
from .filters import BankGuaranteeFilter


class DropdownOptionViewSet(viewsets.ModelViewSet):
    """ViewSet for managing dropdown options."""
    queryset = DropdownOption.objects.all()
    serializer_class = DropdownOptionSerializer
    pagination_class = None

    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        return queryset


class BankGuaranteeViewSet(viewsets.ModelViewSet):
    """ViewSet for Bank Guarantee CRUD operations with import/export and notifications."""
    queryset = BankGuarantee.objects.all()
    serializer_class = BankGuaranteeSerializer
    filterset_class = BankGuaranteeFilter
    ordering_fields = [
        'bank_name', 'branch_name', 'department', 'bg_number',
        'amount_of_bg', 'issue_date', 'expiry_date', 'created_at'
    ]
    ordering = ['-issue_date']

    def get_queryset(self):
        """Default to current year if no year filter is specified."""
        queryset = super().get_queryset()
        # If no year param, and not fetching a single record, default to current year
        if self.action == 'list' and 'year' not in self.request.query_params:
            queryset = queryset.filter(issue_date__year=date.today().year)
        return queryset

    # HEADER MAPPING for flexible import
    HEADER_MAP = {
        'bank_name': ['bank_name', 'bank name', 'bankname', 'bank'],
        'branch_name': ['branch_name', 'branch name', 'branchname', 'branch'],
        'ifsc_code': ['ifsc_code', 'ifsc code', 'ifsccode', 'ifsc'],
        'debit_account_no': ['debit_account_no', 'debit account no', 'debit account number', 'debit acc no', 'debit_acc_no', 'account_no', 'account no', 'accountno', 'account number', 'account_number', 'acc no', 'acc_no'],
        'bg_account_no': ['bg_account_no', 'bg account no', 'bg account number', 'bg acc no', 'bg_acc_no'],
        'payment_mode': ['payment_mode', 'payment mode', 'mode of payment', 'paymentmode'],
        'department': ['department', 'dept', 'dept.'],
        'bg_number': ['bg_number', 'bg number', 'bgnumber', 'bg no', 'bg_no', 'bank guarantee number', 'bank_guarantee_number', 'bank guarantee no'],
        'amount_of_bg': ['amount_of_bg', 'amount of bg', 'amountofbg', 'amount', 'bg amount', 'amount of bank guarantee'],
        'issue_date': ['issue_date', 'issue date', 'issuedate', 'date of issue'],
        'expiry_date': ['expiry_date', 'expiry date', 'expirydate', 'expiry', 'date of expiry'],
        # Cheque fields
        'cheque_number': ['cheque_number', 'cheque number', 'chequenumber', 'cheque no', 'cheque_no'],
        'date_of_issue_of_cheque': ['date_of_issue_of_cheque', 'date of issue of cheque', 'cheque date', 'cheque issue date'],
        'bank_name_of_cheque': ['bank_name_of_cheque', 'bank name of cheque', 'cheque bank', 'cheque bank name'],
        'account_no_of_cheque': ['account_no_of_cheque', 'account no of cheque', 'cheque account', 'cheque account no', 'cheque acc no'],
        # Online fields
        'online_transaction_id': ['online_transaction_id', 'online transaction id', 'transaction id', 'txn id', 'transaction no', 'online_txn_id'],
        'online_transaction_date': ['online_transaction_date', 'online transaction date', 'transaction date', 'txn date', 'online_txn_date'],
        'online_payment_mode': ['online_payment_mode', 'online payment mode', 'online mode', 'online payment channel'],
        'online_bank_name': ['online_bank_name', 'online bank name', 'online bank'],
        # PDC fields
        'pdc_cheque_number': ['pdc_cheque_number', 'pdc cheque number', 'pdc cheque no', 'pdc number', 'pdc no'],
        'pdc_date_of_issue_of_cheque': ['pdc_date_of_issue_of_cheque', 'pdc date of issue of cheque', 'pdc cheque date', 'pdc issue date', 'pdc cheque issue date'],
        'pdc_bank_name_of_cheque': ['pdc_bank_name_of_cheque', 'pdc bank name of cheque', 'pdc bank', 'pdc bank name'],
        'pdc_account_no_of_cheque': ['pdc_account_no_of_cheque', 'pdc account no of cheque', 'pdc cheque account', 'pdc cheque account no', 'pdc cheque acc no'],
    }

    def _normalize_header(self, header):
        """Normalize a header string for flexible matching."""
        return re.sub(r'[^a-z0-9]', ' ', header.lower().strip()).strip()

    def _map_headers(self, headers):
        """Map file headers to model field names using flexible matching."""
        mapping = {}
        for header in headers:
            normalized = self._normalize_header(header)
            for field_name, variations in self.HEADER_MAP.items():
                if normalized in [self._normalize_header(v) for v in variations]:
                    mapping[header] = field_name
                    break
        return mapping

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Import BankGuarantee records from an uploaded Excel or CSV file."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            elif filename.endswith('.csv'):
                import csv
                decoded = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
                headers = reader.fieldnames or []
            else:
                return Response({'error': 'Unsupported file format. Use .xlsx or .csv'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'error': 'No data found in file'}, status=status.HTTP_400_BAD_REQUEST)

        # Map headers flexibly
        header_mapping = self._map_headers(headers)

        created_count = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            mapped_row = {}
            for original_header, value in row.items():
                field_name = header_mapping.get(original_header)
                if field_name:
                    mapped_row[field_name] = value

            # Skip empty rows
            if not any(mapped_row.values()):
                continue

            try:
                # Convert types
                if 'amount_of_bg' in mapped_row and mapped_row['amount_of_bg']:
                    mapped_row['amount_of_bg'] = Decimal(str(mapped_row['amount_of_bg']))

                # Handle date fields
                for date_field in ['issue_date', 'expiry_date', 'date_of_issue_of_cheque', 'online_transaction_date', 'pdc_date_of_issue_of_cheque']:
                    if date_field in mapped_row and mapped_row[date_field]:
                        val = mapped_row[date_field]
                        if hasattr(val, 'date'):
                            mapped_row[date_field] = val.date() if hasattr(val, 'date') else val
                        elif isinstance(val, str):
                            from django.utils.dateparse import parse_date
                            parsed = parse_date(val)
                            if parsed:
                                mapped_row[date_field] = parsed
                            else:
                                # Try common formats
                                from datetime import datetime
                                for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y', '%Y/%m/%d']:
                                    try:
                                        mapped_row[date_field] = datetime.strptime(val, fmt).date()
                                        break
                                    except ValueError:
                                        continue
                    elif date_field in ['issue_date', 'expiry_date']:
                        pass  # Required fields, will be caught by serializer

                # Handle nullable fields
                for nullable in [
                    'cheque_number', 'date_of_issue_of_cheque', 'bank_name_of_cheque', 'account_no_of_cheque',
                    'online_transaction_id', 'online_transaction_date', 'online_payment_mode', 'online_bank_name',
                    'pdc_cheque_number', 'pdc_date_of_issue_of_cheque', 'pdc_bank_name_of_cheque', 'pdc_account_no_of_cheque'
                ]:
                    if nullable in mapped_row and (mapped_row[nullable] is None or str(mapped_row[nullable]).strip() == ''):
                        mapped_row[nullable] = None

                # Adjust based on payment mode
                if 'payment_mode' not in mapped_row or not mapped_row['payment_mode']:
                    if mapped_row.get('online_transaction_id') or mapped_row.get('online_bank_name'):
                        mapped_row['payment_mode'] = 'online'
                    else:
                        mapped_row['payment_mode'] = 'cheque'

                if mapped_row.get('payment_mode') == 'online':
                    for f in ['cheque_number', 'date_of_issue_of_cheque', 'bank_name_of_cheque', 'account_no_of_cheque']:
                        mapped_row[f] = None
                elif mapped_row.get('payment_mode') == 'cheque':
                    for f in ['online_transaction_id', 'online_transaction_date', 'online_payment_mode', 'online_bank_name']:
                        mapped_row[f] = None

                serializer = BankGuaranteeSerializer(data=mapped_row)
                if serializer.is_valid():
                    serializer.save()
                    created_count += 1
                else:
                    errors.append({'row': i, 'errors': serializer.errors})
            except Exception as e:
                errors.append({'row': i, 'errors': str(e)})

        return Response({
            'imported': created_count,
            'errors': errors,
            'total_rows': len(rows),
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export')
    def export_data(self, request):
        """Export filtered BankGuarantee records as an Excel file."""
        # Apply filters to queryset
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Bank Guarantees'

        # Header row
        headers = [
            'Bank Name', 'Branch Name', 'IFSC Code', 'Debit Account No.', 'BG Account No.', 'Payment Mode', 'Department',
            'BG Number', 'Amount of BG', 'PDC', 'Total Amount', 'Quantity (kg)',
            'Issue Date', 'Expiry Date', 'No. of Days',
            'Cheque Number', 'Date of Issue of Cheque', 'Bank Name of Cheque', 'Account No. of Cheque',
            'Online Txn ID', 'Online Txn Date', 'Online Payment Mode', 'Online Bank Name',
            'PDC Cheque Number', 'PDC Cheque Issue Date', 'PDC Bank Name', 'PDC Cheque Account No.'
        ]
        ws.append(headers)

        # Style header row (bold)
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows
        for bg in queryset:
            ws.append([
                bg.bank_name, bg.branch_name, bg.ifsc_code, bg.debit_account_no, bg.bg_account_no or '', bg.payment_mode, bg.department,
                bg.bg_number, float(bg.amount_of_bg), float(bg.pdc), float(bg.total_amount), bg.quantity,
                bg.issue_date.isoformat() if bg.issue_date else '',
                bg.expiry_date.isoformat() if bg.expiry_date else '',
                bg.no_of_days,
                bg.cheque_number or '',
                bg.date_of_issue_of_cheque.isoformat() if bg.date_of_issue_of_cheque else '',
                bg.bank_name_of_cheque or '',
                bg.account_no_of_cheque or '',
                bg.online_transaction_id or '',
                bg.online_transaction_date.isoformat() if bg.online_transaction_date else '',
                bg.online_payment_mode or '',
                bg.online_bank_name or '',
                bg.pdc_cheque_number or '',
                bg.pdc_date_of_issue_of_cheque.isoformat() if bg.pdc_date_of_issue_of_cheque else '',
                bg.pdc_bank_name_of_cheque or '',
                bg.pdc_account_no_of_cheque or '',
            ])

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Write to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=bank_guarantees.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='expiring-soon')
    def expiring_soon(self, request):
        """Return BankGuarantee records expiring within the next 30 days."""
        today = date.today()
        thirty_days = today + timedelta(days=30)
        queryset = BankGuarantee.objects.filter(
            expiry_date__gte=today,
            expiry_date__lte=thirty_days
        ).order_by('expiry_date')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
