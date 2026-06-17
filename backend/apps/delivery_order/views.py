import io
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated

from apps.bank_guarantee.models import DropdownOption
from .models import DeliveryOrder, KaantaParchi, KaantaParchiDOAllocation
from .serializers import DeliveryOrderSerializer, KaantaParchiSerializer
from .filters import DeliveryOrderFilter, KaantaParchiFilter


class DeliveryOrderViewSet(viewsets.ModelViewSet):
    queryset = DeliveryOrder.objects.all()
    serializer_class = DeliveryOrderSerializer
    filterset_class = DeliveryOrderFilter
    permission_classes = [IsAuthenticated]

    def _register_do_location(self, do_location):
        """Register the DO location in DropdownOption if it does not already exist."""
        if do_location:
            val = do_location.strip()
            from apps.bank_guarantee.models import DropdownOption
            if not DropdownOption.objects.filter(category='do_location', value__iexact=val).exists():
                DropdownOption.objects.create(category='do_location', value=val)

    def perform_create(self, serializer):
        instance = serializer.save()
        self._register_do_location(instance.do_location)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._register_do_location(instance.do_location)

    @action(detail=False, methods=['get'], url_path='aggregate-bg-quantity')
    def aggregate_bg_quantity(self, request):
        """Returns the aggregate total quantity figures pulled from the Bank Guarantee module."""
        try:
            from apps.bank_guarantee.models import BankGuarantee
            total_qty = sum(bg.quantity for bg in BankGuarantee.objects.all())
            return Response({"aggregate_bg_quantity": total_qty}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class KaantaParchiViewSet(viewsets.ModelViewSet):
    queryset = KaantaParchi.objects.all()
    serializer_class = KaantaParchiSerializer
    filterset_class = KaantaParchiFilter
    permission_classes = [IsAuthenticated]

    HEADER_MAP = {
        'kaanta_parchi_no': ['kaanta parchi no', 'kaanta parchi number', 'kaanta_parchi_no', 'slip no', 'slip number', 'parchi no'],
        'vehicle_no': ['vehicle no', 'vehicle number', 'vehicle_no', 'truck no', 'truck number'],
        'driver_name': ['driver name', 'driver_name', 'driver'],
        'driver_mobile_no': ['driver mobile', 'driver mobile no', 'driver_mobile_no', 'mobile', 'mobile no', 'phone'],
        'gate_pass_no': ['gate pass no', 'gate pass number', 'gate_pass_no', 'gate pass'],
        'gate_pass_date': ['gate pass date', 'gate_pass_date', 'date'],
        'no_of_boras': ['no of boras', 'no_of_boras', 'boras', 'bags', 'number of boras'],
        'weight_of_empty_truck': ['empty truck weight', 'weight of empty truck', 'weight_of_empty_truck', 'empty weight'],
        'weight_of_filled_truck': ['filled truck weight', 'weight of filled truck', 'weight_of_filled_truck', 'filled weight'],
        'do_number': ['do number', 'do_number', 'do no', 'delivery order', 'do_no'],
    }

    def _register_vehicle_no(self, vehicle_no):
        """Register the vehicle number in DropdownOption if it does not already exist."""
        if vehicle_no:
            val = vehicle_no.strip()
            if not DropdownOption.objects.filter(category='vehicle_no', value__iexact=val).exists():
                DropdownOption.objects.create(category='vehicle_no', value=val)

    def perform_create(self, serializer):
        instance = serializer.save()
        self._register_vehicle_no(instance.vehicle_no)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._register_vehicle_no(instance.vehicle_no)

    def _normalize_header(self, header):
        return str(header or '').strip().lower().replace('_', ' ').replace('-', ' ')

    def _map_headers(self, headers):
        mapping = {}
        for header in headers:
            normalized = self._normalize_header(header)
            for field_name, variations in self.HEADER_MAP.items():
                if normalized in [self._normalize_header(v) for v in variations]:
                    mapping[header] = field_name
                    break
        return mapping

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Bulk import KaantaParchi records from an Excel or CSV file."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Skip completely empty rows
                        rows.append(dict(zip(headers, row)))
            elif filename.endswith('.csv'):
                import csv
                decoded = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
                headers = reader.fieldnames or []
            else:
                return Response({'error': 'Unsupported file format. Use .xlsx or .csv'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'error': 'No data found in file'}, status=status.HTTP_400_BAD_REQUEST)

        header_mapping = self._map_headers(headers)
        created_count = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            mapped_row = {}
            for original_header, value in row.items():
                field_name = header_mapping.get(original_header)
                if field_name:
                    mapped_row[field_name] = value

            if not any(mapped_row.values()):
                continue

            try:
                # Type Conversions
                if 'no_of_boras' in mapped_row and mapped_row['no_of_boras'] is not None:
                    mapped_row['no_of_boras'] = int(float(str(mapped_row['no_of_boras'])))
                
                if 'weight_of_empty_truck' in mapped_row and mapped_row['weight_of_empty_truck'] is not None:
                    mapped_row['weight_of_empty_truck'] = Decimal(str(mapped_row['weight_of_empty_truck']))
                
                if 'weight_of_filled_truck' in mapped_row and mapped_row['weight_of_filled_truck'] is not None:
                    mapped_row['weight_of_filled_truck'] = Decimal(str(mapped_row['weight_of_filled_truck']))

                # Handle date fields
                if 'gate_pass_date' in mapped_row and mapped_row['gate_pass_date']:
                    val = mapped_row['gate_pass_date']
                    if hasattr(val, 'date'):
                        mapped_row['gate_pass_date'] = val.date()
                    elif isinstance(val, str):
                        parsed = parse_date(val)
                        if parsed:
                            mapped_row['gate_pass_date'] = parsed
                        else:
                            # Try common formats
                            for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y', '%Y/%m/%d']:
                                try:
                                    mapped_row['gate_pass_date'] = datetime.strptime(val, fmt).date()
                                    break
                                except ValueError:
                                    continue

                # Retrieve/validate Delivery Order for allocation
                do_number_val = mapped_row.pop('do_number', None)
                if not do_number_val:
                    errors.append({'row': i, 'errors': {'do_number': 'do_number column is required for allocation'}})
                    continue

                do_number_val = str(do_number_val).strip()
                try:
                    do_obj = DeliveryOrder.objects.get(do_number__iexact=do_number_val)
                except DeliveryOrder.DoesNotExist:
                    errors.append({'row': i, 'errors': {'do_number': f'Delivery Order with number "{do_number_val}" does not exist.'}})
                    continue

                # Build nested allocations
                mapped_row['do_allocations'] = [{
                    'delivery_order_id': do_obj.id,
                    'allocated_boras': mapped_row['no_of_boras']
                }]

                serializer = KaantaParchiSerializer(data=mapped_row)
                if serializer.is_valid():
                    instance = serializer.save()
                    self._register_vehicle_no(instance.vehicle_no)
                    created_count += 1
                else:
                    errors.append({'row': i, 'errors': serializer.errors})
            except Exception as e:
                errors.append({'row': i, 'errors': str(e)})

        return Response({
            'imported': created_count,
            'errors': errors,
            'total_rows': len(rows),
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export')
    def export_data(self, request):
        """Export filtered KaantaParchi records as an Excel file."""
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Kaanta Parchis'

        headers = [
            'Kaanta Parchi No.', 'Vehicle No.', 'Driver Name', 'Driver Mobile', 'Gate Pass No.', 'Gate Pass Date',
            'Issued No. of Sacks', 'Sacks Allocated', 'Weight of Boras (kg)', 'Weight of Dhan (kg)', 'Weight of Empty Truck (kg)',
            'Weight of Filled Truck (kg)', 'Net Weight (kg)', 'DO(s) Allocated'
        ]
        ws.append(headers)

        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for kp in queryset:
            allocations_str = ", ".join([
                f"{alloc.delivery_order.do_number} ({alloc.allocated_boras} bags)"
                for alloc in kp.do_allocations.all()
            ])
            ws.append([
                kp.kaanta_parchi_no, kp.vehicle_no, kp.driver_name, kp.driver_mobile_no, kp.gate_pass_no,
                kp.gate_pass_date.isoformat() if kp.gate_pass_date else '',
                kp.no_of_boras, kp.no_of_boras, float(kp.weight_of_boras), float(kp.weight_of_dhan),
                float(kp.weight_of_empty_truck), float(kp.weight_of_filled_truck), float(kp.net_weight),
                allocations_str
            ])

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=kaanta_parchis.xlsx'
        wb.save(response)
        return response
